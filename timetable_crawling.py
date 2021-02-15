from os import execlp
from selenium import webdriver
from bs4 import BeautifulSoup
from openpyxl import Workbook,load_workbook
import time
from selenium.webdriver.common.keys import Keys
import pandas as pd

driver = webdriver.Chrome("./chromedriver")
add = 'https://everytime.kr/timetable/2021/1'
driver.get(add)

#로그인
driver.find_element_by_css_selector('#container > form > p:nth-child(1) > input').send_keys('chdnjf103')
driver.find_element_by_css_selector('#container > form > p:nth-child(2) > input').send_keys('a478956')
driver.find_element_by_css_selector('#container > form > p.submit > input').click()

time.sleep(3)


# 광고 닫기
driver.find_element_by_css_selector('#sheet > ul > li.none > a').click()

time.sleep(2)

# 시간표 목록
driver.find_element_by_css_selector("#container > ul > li.button.search").click()

time.sleep(2)



# 심교만
driver.find_element_by_css_selector('#subjects > div.filter > a:nth-child(9) > span.key').click()
driver.find_element_by_css_selector('#subjectTypeFilter > input:nth-child(5)').click()
driver.find_element_by_css_selector('#subjectTypeFilter > div > label:nth-child(5) > input[type=checkbox]').click() 
driver.find_element_by_css_selector('#subjectTypeFilter > input:nth-child(6)').click() 

time.sleep(3)

#스크롤 맨아래로 내리기
pre = 0
while True:
    element = driver.find_elements_by_css_selector("#subjects > div.list > table > tbody > tr")
    result = element[-1]
    driver.execute_script('arguments[0].scrollIntoView(true);',result)
    time.sleep(2)

    current = len(element)
    if pre == current:
        break
    pre = current


html = driver.page_source
soup = BeautifulSoup(html, 'html.parser')

trs = soup.select('#subjects > div.list > table > tbody > tr')
results = []

for tr in trs:
    result=[]
    tds = tr.select('td')
    result.append(tds[3].text) # 과목번호
    result.append(tds[4].text) # 강의명
    result.append(tds[5].text) # 학점
    result.append(tds[6].text) # 교수님
    result.append(tds[7].text) # 시간
    result.append(tds[12].text) # 영역
    # 별점
    title = tr.find('a', attrs={'class':'star'}) 
    print(title)
    print(title['title'])
    result.append(title['title'])
    
    results.append(result)
    

# 엑셀에 저장
write_wb = Workbook()

for data in results:
    write_ws = write_wb.active
    write_ws.append(data)
write_wb.save('심교.csv')


# 닫기
driver.close()




"""
강의평
"""
driver = webdriver.Chrome("./chromedriver")
add = 'https://everytime.kr/lecture'
driver.get(add)

#로그인
driver.find_element_by_css_selector('#container > form > p:nth-child(1) > input').send_keys('chdnjf103')
driver.find_element_by_css_selector('#container > form > p:nth-child(2) > input').send_keys('a478956')
driver.find_element_by_css_selector('#container > form > p.submit > input').click()

time.sleep(3)


# 광고 닫기
driver.find_element_by_css_selector('#sheet > ul > li.none > a').click()

time.sleep(2)

# 강의 검색
table = pd.read_csv('심교.csv',header=None)
table_1 = table.iloc[:,[1,3]].drop_duplicates(keep='first')

res = []
for name,prof in zip(table_1.iloc[:,0].tolist(),table_1.iloc[:,1].tolist()):
    time.sleep(1)
    driver.find_element_by_css_selector(".keyword").clear()
    driver.find_element_by_css_selector(".keyword").send_keys(name)
    driver.find_element_by_css_selector(".keyword").send_keys(Keys.ENTER)
    time.sleep(2)
    try:
        driver.find_element_by_xpath("//*[contains(text(), prof)]").click()
    except:
        continue

    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')
    time.sleep(2)

    try:
        article_lst = soup.select('.articles')[0].select('.text')
        summary_lst = soup.select('.details')[0].select('span')
    except:
        driver.back()
        continue

    result = [name,prof]
    for j in summary_lst:
        result.append(j.text)

    for i in article_lst:
        result.append(i.text)

    res.append(result)
    driver.back()

res

    

# 엑셀에 저장
write_wb = Workbook()

for data in res:
    write_ws = write_wb.active
    write_ws.append(data)
write_wb.save('강의평.csv')


# 닫기
driver.close()